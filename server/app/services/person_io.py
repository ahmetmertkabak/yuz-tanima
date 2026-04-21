"""
Excel / CSV import + export helpers for Person records.

Import:
    summary = PersonImporter(school_id=1).import_file(file_stream, default_role='student')
    # → {"created": 40, "updated": 3, "errors": [...]}

Export:
    buf = PersonExporter.persons_to_excel(queryset)
    send_file(buf, mimetype=..., as_attachment=True, download_name="persons.xlsx")
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Iterable, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.extensions import db
from app.models import ConsentStatus, Person, PersonRole


# Column aliases (lower-cased) → canonical field name
COLUMN_ALIASES: dict[str, str] = {
    # person_no
    "person_no": "person_no",
    "no": "person_no",
    "sicil": "person_no",
    "ogrenci_no": "person_no",
    "öğrenci no": "person_no",
    "öğrenci_no": "person_no",
    # full_name
    "full_name": "full_name",
    "ad_soyad": "full_name",
    "ad soyad": "full_name",
    "isim": "full_name",
    "name": "full_name",
    # role
    "role": "role",
    "rol": "role",
    "tip": "role",
    # class_name
    "class_name": "class_name",
    "sinif": "class_name",
    "sınıf": "class_name",
    "class": "class_name",
    # email
    "email": "email",
    "e-posta": "email",
    "eposta": "email",
    # phone
    "phone": "phone",
    "telefon": "phone",
    "tel": "phone",
    # parent
    "parent_phone": "parent_phone",
    "veli_telefon": "parent_phone",
    "veli telefon": "parent_phone",
    "parent_name": "parent_name",
    "veli_adi": "parent_name",
    "veli adı": "parent_name"}

ROLE_ALIASES: dict[str, str] = {
    "öğrenci": PersonRole.STUDENT.value,
    "ogrenci": PersonRole.STUDENT.value,
    "student": PersonRole.STUDENT.value,
    "teacher": PersonRole.TEACHER.value,
    "öğretmen": PersonRole.TEACHER.value,
    "ogretmen": PersonRole.TEACHER.value,
    "staff": PersonRole.STAFF.value,
    "personel": PersonRole.STAFF.value,
    "manager": PersonRole.MANAGER.value,
    "yönetici": PersonRole.MANAGER.value,
    "yonetici": PersonRole.MANAGER.value}


@dataclass
class ImportRowError:
    row: int
    field: str
    message: str
    raw: dict


@dataclass
class ImportSummary:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[ImportRowError] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": [e.__dict__ for e in self.errors],
            "total": self.created + self.updated + self.skipped + len(self.errors)}


# ---------------------------------------------------------------------------
# Importer
# ---------------------------------------------------------------------------
class PersonImporter:
    """Imports rows into `persons` for a specific school."""

    def __init__(
        self,
        school_id: int,
        default_role: str = PersonRole.STUDENT.value,
        overwrite: bool = False,
    ) -> None:
        self.school_id = school_id
        self.default_role = default_role
        self.overwrite = overwrite

    # ---- public ----
    def import_file(self, stream, filename: str = "") -> ImportSummary:
        rows = self._read_rows(stream, filename)
        return self._import_rows(rows)

    # ---- internals ----
    def _read_rows(self, stream, filename: str) -> list[dict]:
        name = (filename or "").lower()
        if name.endswith(".csv"):
            return self._read_csv(stream)
        return self._read_xlsx(stream)

    def _read_xlsx(self, stream) -> list[dict]:
        wb = openpyxl.load_workbook(stream, data_only=True)
        ws = wb.active
        if ws.max_row < 2:
            return []

        header_raw = [c.value for c in ws[1]]
        header = [self._canonical_column(h) for h in header_raw]

        rows: list[dict] = []
        for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = {}
            for col_name, value in zip(header, raw):
                if col_name and value not in (None, ""):
                    row[col_name] = str(value).strip()
            if row:
                row["__row__"] = idx
                rows.append(row)
        return rows

    def _read_csv(self, stream) -> list[dict]:
        import csv

        try:
            text = stream.read().decode("utf-8-sig")
        except AttributeError:
            text = stream.read()
        reader = csv.DictReader(io.StringIO(text))
        rows: list[dict] = []
        for idx, raw in enumerate(reader, start=2):
            row = {
                self._canonical_column(k): (v or "").strip()
                for k, v in raw.items()
                if k and v
            }
            row = {k: v for k, v in row.items() if k}
            if row:
                row["__row__"] = idx
                rows.append(row)
        return rows

    @staticmethod
    def _canonical_column(name) -> Optional[str]:
        if not name:
            return None
        key = str(name).strip().lower().replace("\n", " ")
        return COLUMN_ALIASES.get(key)

    def _import_rows(self, rows: list[dict]) -> ImportSummary:
        summary = ImportSummary()

        for row in rows:
            row_no = row.pop("__row__", 0)

            person_no = (row.get("person_no") or "").strip()
            full_name = (row.get("full_name") or "").strip()

            if not person_no:
                summary.errors.append(
                    ImportRowError(row_no, "person_no", "Zorunlu alan eksik.", row)
                )
                continue
            if not full_name:
                summary.errors.append(
                    ImportRowError(row_no, "full_name", "Zorunlu alan eksik.", row)
                )
                continue

            role = self._normalize_role(row.get("role")) or self.default_role
            if role not in PersonRole.values():
                summary.errors.append(
                    ImportRowError(row_no, "role", f"Geçersiz rol: {row.get('role')}", row)
                )
                continue

            existing = (
                db.session.query(Person)
                .filter_by(school_id=self.school_id, person_no=person_no)
                .first()
            )

            if existing and not self.overwrite:
                summary.skipped += 1
                continue

            data = {
                "school_id": self.school_id,
                "person_no": person_no,
                "full_name": full_name,
                "role": role,
                "class_name": row.get("class_name") or None,
                "email": row.get("email") or None,
                "phone": row.get("phone") or None,
                "parent_phone": row.get("parent_phone") or None,
                "parent_name": row.get("parent_name") or None,
                "is_active": True,
                "access_granted": True}

            try:
                if existing:
                    for k, v in data.items():
                        setattr(existing, k, v)
                    summary.updated += 1
                else:
                    # New persons default to pending consent
                    data["consent_status"] = ConsentStatus.PENDING.value
                    db.session.add(Person(**data))
                    summary.created += 1
            except Exception as exc:  # pragma: no cover
                summary.errors.append(
                    ImportRowError(row_no, "_db", str(exc), row)
                )

        return summary

    @staticmethod
    def _normalize_role(value) -> Optional[str]:
        if not value:
            return None
        key = str(value).strip().lower()
        return ROLE_ALIASES.get(key, key if key in PersonRole.values() else None)


# ---------------------------------------------------------------------------
# Exporter
# ---------------------------------------------------------------------------
class PersonExporter:
    """Generates an Excel file from a Person queryset/iterable."""

    HEADER = [
        ("person_no", "No"),
        ("full_name", "Ad Soyad"),
        ("role", "Rol"),
        ("class_name", "Sınıf"),
        ("email", "E-posta"),
        ("phone", "Telefon"),
        ("parent_name", "Veli Adı"),
        ("parent_phone", "Veli Telefonu"),
        ("consent_status", "KVKK"),
        ("is_active", "Aktif"),
        ("has_face", "Yüz Kayıtlı"),
        ("created_at", "Oluşturulma")]

    @classmethod
    def persons_to_excel(cls, persons: Iterable[Person]) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kişiler"

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)

        ws.append([label for _, label in cls.HEADER])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        for p in persons:
            ws.append(
                [
                    p.person_no,
                    p.full_name,
                    p.role,
                    p.class_name or "",
                    p.email or "",
                    p.phone or "",
                    p.parent_name or "",
                    p.parent_phone or "",
                    p.consent_status,
                    "Evet" if p.is_active else "Hayır",
                    "Evet" if p.has_face else "Hayır",
                    p.created_at.strftime("%Y-%m-%d") if p.created_at else ""]
            )

        # Auto column widths
        widths = [12, 28, 12, 12, 28, 18, 22, 18, 12, 8, 12, 12]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf