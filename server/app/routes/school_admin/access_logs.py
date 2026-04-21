"""Access log browsing + export."""
from __future__ import annotations

from datetime import datetime, timedelta

from flask import render_template, request, send_file
from flask_login import current_user, login_required

from app.extensions import db
from app.middleware.auth import school_staff_required
from app.middleware.tenant import get_tenant_context, require_school_context
from app.models import (
    AccessLog,
    AccessOutcome,
    AuditAction,
    Device,
    Person,
)
from app.routes.school_admin import bp
from app.services.audit import record_audit


def _parse_date(value: str | None, default: datetime) -> datetime:
    if not value:
        return default
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return default


def _logs_query():
    now = datetime.utcnow()
    default_from = datetime(now.year, now.month, now.day)
    default_to = default_from + timedelta(days=1)

    date_from = _parse_date(request.args.get("from"), default_from)
    date_to = _parse_date(request.args.get("to"), default_to) + timedelta(days=1)
    # +1 day because the `to` date is inclusive

    person_id = request.args.get("person_id", type=int)
    device_id = request.args.get("device_id", type=int)
    outcome = request.args.get("outcome") or ""
    denied_only = request.args.get("denied") == "1"

    query = (
        db.session.query(AccessLog)
        .filter(
            AccessLog.event_at >= date_from,
            AccessLog.event_at < date_to,
        )
        .order_by(AccessLog.event_at.desc())
    )
    if person_id:
        query = query.filter(AccessLog.person_id == person_id)
    if device_id:
        query = query.filter(AccessLog.device_id == device_id)
    if outcome and outcome in AccessOutcome.values():
        query = query.filter(AccessLog.outcome == outcome)
    if denied_only:
        query = query.filter(AccessLog.outcome != AccessOutcome.GRANTED.value)

    return query, {
        "from": date_from.strftime("%Y-%m-%d"),
        "to": (date_to - timedelta(days=1)).strftime("%Y-%m-%d"),
        "person_id": person_id,
        "device_id": device_id,
        "outcome": outcome,
        "denied_only": denied_only}


@bp.route("/access-logs")
@login_required
@require_school_context
@school_staff_required
def access_logs_list():
    query, filters = _logs_query()

    page = max(int(request.args.get("page") or 1), 1)
    per_page = min(int(request.args.get("per_page") or 50), 200)

    total = query.count()
    logs = query.offset((page - 1) * per_page).limit(per_page).all()

    devices = db.session.query(Device).order_by(Device.device_name).all()
    outcomes = AccessOutcome.values()

    return render_template(
        "school_admin/access_logs/list.html",
        logs=logs,
        total=total,
        page=page,
        per_page=per_page,
        filters=filters,
        devices=devices,
        outcomes=outcomes,
    )


@bp.route("/access-logs/export")
@login_required
@require_school_context
@school_staff_required
def access_logs_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    query, filters = _logs_query()
    logs = query.limit(50_000).all()  # safety cap

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Geçişler"
    headers = ["Zaman", "Ad Soyad", "No", "Sınıf", "Kapı", "Yön", "Sonuç", "Güven"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(color="FFFFFF", bold=True)

    for log in logs:
        ws.append(
            [
                log.event_at.strftime("%Y-%m-%d %H:%M:%S"),
                log.person_name_cached or "Bilinmeyen",
                log.person_no_cached or "",
                log.person_class_cached or "",
                log.device.device_name if log.device else "",
                log.direction,
                log.outcome,
                f"{log.confidence:.2f}" if log.confidence else ""]
        )

    widths = [20, 26, 14, 12, 20, 10, 18, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ctx = get_tenant_context()
    record_audit(
        AuditAction.DATA_EXPORT_REQUESTED,
        user=current_user,
        school_id=ctx.school_id,
        resource_type="access_logs",
        resource_label=f"{len(logs)} satır",
        details=filters,
        commit=True,
    )

    return send_file(
        buf,
        as_attachment=True,
        download_name=(
            f"geciler-{ctx.school.subdomain}-{filters['from']}-{filters['to']}.xlsx"
        ),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )